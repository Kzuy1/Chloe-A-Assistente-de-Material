import ezdxf
import openpyxl

workbook = openpyxl.load_workbook('material1.xlsx', data_only= True)
worksheet = workbook['Sheet1']

PecaList = []

for row in range(2, worksheet.max_row + 1):
  cell_value = worksheet.cell(row=row, column=1).value
  if cell_value == None:
    break
  peca = {
    'cod' : worksheet.cell(row=row, column=1).value,
    'descricao' : worksheet.cell(row=row, column=2).value,
    'unidade' : worksheet.cell(row=row, column=3).value,
    'QTD' : worksheet.cell(row=row, column=4).value,
    'material' : worksheet.cell(row=row, column=5).value,
    'peso' : worksheet.cell(row=row, column=6).value
  }
  PecaList.append(peca)

doc = ezdxf.readfile('Drawing3.dxf')

bloco_original = doc.blocks.get('EMB_LISTA_DE_MATERIAL')

msp = doc.modelspace()

if bloco_original :
  for indice, objeto in enumerate(PecaList):
      pos = 10 + 7 * indice
      insert = msp.add_blockref(bloco_original.name, insert=(0, pos, 0))
      insert.add_attrib("POSICAO", objeto["cod"])
      insert.add_attrib("DESCRICAO", objeto["descricao"])
      insert.add_attrib("UNIDADE", objeto["unidade"])
      insert.add_attrib("QUANTIDADE", objeto["QTD"])
      insert.add_attrib("MATERIAL", objeto["material"])
      insert.add_attrib("PESO", objeto["peso"])


doc.saveas("BLOCO-BRANCO-GRADE_updated.dxf")

