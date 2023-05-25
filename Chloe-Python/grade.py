import ezdxf
import openpyxl

workbook = openpyxl.load_workbook('2023-05-10_GRADES_SK-B9-22.xlsx', data_only= True)
worksheet = workbook['BLOCO BRANCO']

PecaList = []

for row in range(3, worksheet.max_row + 1):
  cell_value = worksheet.cell(row=row, column=5).value
  if cell_value == None:
    break
  peca = {
    'cod' : worksheet.cell(row=row, column=1).value,
    'descricao' : worksheet.cell(row=row, column=2).value,
    'QTDE' : worksheet.cell(row=row, column=3).value,
    'material' : worksheet.cell(row=1, column=4).value,
    'PesoUnit' : worksheet.cell(row=row, column=4).value,
    'PesoTotal' : worksheet.cell(row=row, column=5).value
  }
  PecaList.append(peca)

doc = ezdxf.readfile('BLOCO-BRANCO-GRADE.dxf')

bloco_original = doc.blocks.get('REDECAM-DISTINTA_monolingua')

msp = doc.modelspace()

if bloco_original :
  for indice, objeto in enumerate(PecaList):
      pos = 10 + 6 * indice
      insert = msp.add_blockref(bloco_original.name, insert=(0, pos, 0))
      insert.add_attrib("MARCA", objeto["cod"])
      insert.add_attrib("DESCRIZIONE-IT", objeto["descricao"])
      insert.add_attrib("DESCRIZIONE-IN-R1", objeto["descricao"])
      insert.add_attrib("QUANTITA'", objeto["QTDE"])
      insert.add_attrib("MATERIALE", objeto["material"])
      insert.add_attrib("PESO-CAD", objeto["PesoUnit"])
      insert.add_attrib("TOTALE", objeto["PesoTotal"])


doc.saveas("BLOCO-BRANCO-GRADE_updated.dxf")

