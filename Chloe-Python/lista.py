import ezdxf
import openpyxl

workbook = openpyxl.load_workbook('./Chloe-Python/2023-05-17_LISTA DE MATERIAL_PLATAFORMA AR FALSO_00_CHLOE_LES.xlsx', data_only= True)

def writeFiles(elemento):
  worksheetMaterial = workbook[elemento + ' - MATERIAL']
  MaterialList = []

  for row in range(2, worksheetMaterial.max_row + 1):
    cell_value = worksheetMaterial.cell(row=row, column=12).value
    if cell_value == None:
      break
    material = {
      'POS' : worksheetMaterial.cell(row=row, column=12).value,
      'descricao' : worksheetMaterial.cell(row=row, column=13).value,
      'unidade' : worksheetMaterial.cell(row=row, column=14).value,
      'QTDE' : worksheetMaterial.cell(row=row, column=15).value,
      'material' : worksheetMaterial.cell(row=row, column=16).value,
      'PesoTotal' : worksheetMaterial.cell(row=row, column=17).value
    }
    MaterialList.append(material)

  worksheetPeca = workbook[elemento + ' - PEÇAS']
  PecaList = []

  for row in range(2, worksheetPeca.max_row + 1):
    cell_value = worksheetMaterial.cell(row=row, column=1).value
    if cell_value == None:
      break
    peca = {
      'cod' : '...' + worksheetPeca.cell(row=row, column=2).value,
      'descricao' : worksheetPeca.cell(row=row, column=6).value or worksheetPeca.cell(row=row, column=7).value or worksheetPeca.cell(row=row, column=5).value,
      'QTDE' : worksheetPeca.cell(row=row, column=4).value,
      'material' : worksheetPeca.cell(row=row, column=8).value,
      'PesoUnit' : worksheetPeca.cell(row=row, column=11).value,
      'PesoTotal' : worksheetPeca.cell(row=row, column=12).value
    }
    PecaList.append(peca)

  doc = ezdxf.readfile('./Chloe-Python/REDECAM-BLOCO-BRANCO.dxf')
  msp = doc.modelspace()
  bloco_material = doc.blocks.get('EMB_LISTA_DE_MATERIAL')

  if bloco_material :
    for indice, objeto in enumerate(MaterialList):
        pos = 196 + 7 * indice
        insert = msp.add_blockref(bloco_material.name, insert=(0, pos, 0))
        insert.add_attrib("POSICAO", objeto["POS"])
        insert.add_attrib("DESCRICAO", objeto["descricao"])
        insert.add_attrib("UNIDADE", objeto["unidade"])
        insert.add_attrib("QUANTIDADE", objeto["QTDE"])
        insert.add_attrib("MATERIAL", objeto["material"])
        insert.add_attrib("PESO", objeto["PesoTotal"])

  bloco_original = doc.blocks.get('REDECAM-DISTINTA_monolingua')

  if bloco_original :
    for indice, objeto in enumerate(PecaList):
        pos = 293 + 6 * indice
        insert = msp.add_blockref(bloco_original.name, insert=(0, pos, 0))
        insert.add_attrib("MARCA", objeto["cod"])
        insert.add_attrib("DESCRIZIONE-IT", objeto["descricao"])
        insert.add_attrib("DESCRIZIONE-IN-R1", objeto["descricao"])
        insert.add_attrib("QUANTITA'", objeto["QTDE"])
        insert.add_attrib("MATERIALE", objeto["material"])
        insert.add_attrib("PESO-CAD", objeto["PesoUnit"])
        insert.add_attrib("TOTALE", objeto["PesoTotal"])

  doc.saveas("./Chloe-Python/" + elemento + "_LISTA.dxf")

planilhas = workbook.worksheets
nomes_planilhas = set()

for planilha in planilhas[1:]:
    primeiro_valor = planilha.title.split('-')[0].strip()
    nomes_planilhas.add(primeiro_valor)

nomes_planilhas = list(nomes_planilhas)
nomes_planilhas.sort()

for elemento in nomes_planilhas:
   writeFiles(elemento)
